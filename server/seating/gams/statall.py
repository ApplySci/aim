"""
Seating Statistics Generator and Analyzer

This module provides functionality to generate and analyze statistics for
various seating configurations in a tournament or game setting. It processes
different player and hanchan (round) combinations, calculates various
statistics, and optionally writes the results to files.

The main functions in this module are:
- stat_all: Generate and analyze statistics for all seating configurations
- write_stats_to_csv: Write general statistical data to a CSV file
- write_meet_stats_to_csv: Write meeting statistics and repeat counts to a CSV file

This module depends on the 'make_stats' function from an external module,
and uses the 'tabulate' library for formatting output.

Usage:
    To run the full analysis and generate all statistics:
    python statall.py

Note: This script assumes the existence of a 'final' directory containing
seating configuration files named in the format 'NxM.py', where N is the
number of players and M is the number of hanchan.
"""

from collections import Counter
import csv
import importlib
import os
import re
from typing import Dict, List, Tuple, Union, Optional

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import tabulate

from make_stats import make_stats

MAX_HANCHAN = 15

def stat_all(write_final: bool = False) -> None:
    """
    Generate and analyze statistics for various seating configurations.

    This function processes different player and hanchan combinations,
    calculates statistics, and optionally writes the results to files.

    Args:
        write_final (bool): If True, write the final seating arrangements to files.

    Returns:
        None
    """
    current_hanchan: Optional[int] = None
    worksheet = None

    pattern = re.compile(r'(\d+)x(\d+)\.py$')
    dimensions: List[Tuple[str, str]] = [
        pattern.match(f).groups()
        for f in os.listdir('final/')
        if pattern.match(f)
    ]
    
    sorted_dimensions: List[Tuple[int, int]] = sorted(
        [(int(players), int(hanchan)) for players, hanchan in dimensions],
        key=lambda x: (x[1], x[0])
    )

    def write_dict(hanchan_count: int,
                   seats_dict: Dict[int, List[List[int]]],
                   warnings: List[List[Union[int, str]]]) -> None:
        """
        Write seating arrangements and warnings to a file.

        Args:
            hanchan_count (int): Number of hanchan.
            seats_dict (Dict[int, List[List[int]]]): Dictionary of seating arrangements.
            warnings (List[List[Union[int, str]]]): List of warnings for imperfect allocations.

        Returns:
            None
        """
        if not write_final or hanchan_count > MAX_HANCHAN:
            return
        out_file = f"../seats_{hanchan_count}.py"
        warnings_table = tabulate.tabulate(
            warnings,
            headers=['players', 'total', 'wind', 'table', 'meet', 'indirect'],
            tablefmt="psql"
        )
        with open(out_file, "w") as f:
            f.write(
                f"""\
\"\"\"
*** imperfect allocations:
{warnings_table}
***
\"\"\"

seats = {{
"""
            )
        
            for player_count, seats in seats_dict.items():
                seats_str = str(seats).replace("],", "],\n    ")
                f.write(f"    {player_count}: {seats_str},\n\n")
            f.write("}\n")
            # end of write_dict


    def end_of_hanchan():
        write_dict(current_hanchan, current_dict, all_warnings)
        # After populating the worksheet, set column widths to autofit
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.4
            worksheet.column_dimensions[column_letter].width = adjusted_width


    # New dictionaries to store statistics for all configurations
    all_wind_hist: Dict[str, Counter] = {}
    all_table_hist: Dict[str, Counter] = {}
    all_meet_hist: Dict[str, Counter] = {}
    all_indirect_hist: Dict[str, Counter] = {}
    all_repeat3: Dict[str, int] = {}
    all_repeat4: Dict[str, int] = {}
    workbook = openpyxl.Workbook()

    for player_count, hanchan_count in sorted_dimensions:
        if hanchan_count != current_hanchan:
            if current_hanchan is not None:
                end_of_hanchan()
            # new output file, so reset all of the variables we accumulate stuff in
            current_hanchan = hanchan_count
            current_dict = {}
            wind_min = hanchan_count // 4
            wind_max = (hanchan_count + 3) // 4
            meet_max = 1
            total_warnings = {}
            wind_warnings = {}
            table_warnings = {}
            meet_warnings = {}
            indirect_warnings = {}
            all_warnings = []
            if hanchan_count <= MAX_HANCHAN:
                worksheet = workbook.create_sheet(title=f"{hanchan_count} h")
            row = 2

        if hanchan_count > MAX_HANCHAN:
            continue
        table_count = player_count // 4
        # take square root of player count to get a reasonable minimum for indirect
        fn = f"{player_count}x{hanchan_count}"
        seats = importlib.import_module(f"final.{fn}").seats

        cell = worksheet.cell(row=row, column=1, value=f"Players: {player_count}")
        cell.font = Font(bold=True)

        for h in range(hanchan_count):
            row += 1
            cell = worksheet.cell(row=row, column=1, value=f"Hanchan {h+1}")
            cell.font = Font(bold=True)
            for t in range(table_count):
                cell = worksheet.cell(row=1, column=t*5+2, value=f"Table {t+1}")
                cell.font = Font(bold=True)
                for p in range(4):
                    worksheet.cell(row=row, column=t*5+p+2, value=seats[h][t][p])
        row += 2
            
        #   meetups. Don't worry about indirect meetups for 3 or fewer hanchan
        indirect_min = int(player_count ** 0.5) if hanchan_count >= 4 else 0
        table_max = (hanchan_count + table_count * 2 - 1) // table_count
        current_dict[player_count] = seats
        stats = make_stats(seats)
        meet_hist, indirect_hist, table_hist, wind_hist = (
            stats['meets'],
            stats['indirects'],
            stats['tables'],
            stats['wind']
        )
        repeat3, repeat4 = stats['repeat3'], stats['repeat4']

        # Store statistics for this configuration
        all_wind_hist[fn] = wind_hist
        all_table_hist[fn] = table_hist
        all_meet_hist[fn] = meet_hist
        all_indirect_hist[fn] = indirect_hist
        all_repeat3[fn] = repeat3
        all_repeat4[fn] = repeat4

        warnings = 0
        wind_warnings = 0
        for n in range(wind_min):
            if wind_hist[n] > 0:
                wind_warnings += 1

        for n in range(wind_max + 1, min(len(wind_hist), hanchan_count + 1)):
            if wind_hist[n] > 0:
                warnings += 1

        table_warnings = 0
        for n in range(table_max + 1, min(len(table_hist), hanchan_count + 1)):
            if table_hist[n] > 0:
                table_warnings += 1

        meet_warnings = 0
        for n in range(meet_max + 1, min(len(meet_hist), hanchan_count + 1)):
            if meet_hist[n] > 0:
                meet_warnings += 1

        indirect_warnings = 0
        for n in range(0, min(len(indirect_hist), indirect_min)):
            if indirect_hist[n] > 0:
                indirect_warnings += 1

        total_warnings = (wind_warnings + table_warnings + meet_warnings +
                          indirect_warnings)
        print(f"{fn}: {total_warnings}")

        wind_warning_text = (
            f"{wind_warnings}/{4*player_count}"
            if wind_warnings > 0 else ''
        )
        table_warning_text = (
            f"{table_warnings}/{player_count * hanchan_count // 2}"
            if table_warnings > 0 else ''
        )
        meet_warning_text = (
            f"{meet_warnings}/{player_count * 3 * hanchan_count // 2}"
            if meet_warnings > 0 else ''
        )
        indirect_warning_text = (
            f"{indirect_warnings}/{player_count * (player_count-1) // 2}"
            if indirect_warnings > 0 else ''
        )

        all_warnings.append([
            player_count, total_warnings, wind_warning_text,
            table_warning_text, meet_warning_text, indirect_warning_text
        ])

    end_of_hanchan()

    stats_data = [
        ("Meets", all_meet_hist, all_repeat3, all_repeat4),
        ("Winds", all_wind_hist),
        ("Tables", all_table_hist),
        ("Indirects", all_indirect_hist)
    ]

    for sheet_name, *data in stats_data:
        ws = workbook.create_sheet(title=sheet_name)
        if sheet_name == "Meets":
            write_meet_stats_to_worksheet(ws, *data)
        else:
            write_stats_to_worksheet(ws, data[0])

    readme_ws = workbook['Sheet']
    readme_ws.title = "README"
    readme_ws.column_dimensions['A'].width = 100
    readme_ws.cell(row=1, column=1,
        value="Seating arrangements for a specific number of hanchan are all given on a single tab."
        )
    readme_ws.cell(row=2, column=1,
        value="e.g. all the seating for 5-hanchan tournaments is given on tab 5 h"
        )
    readme_ws.cell(row=4, column=1,
        value="Player IDs are a 1-based sequence: e.g. for a 48-player tournament, player IDs run 1-48"
        )
    readme_ws.cell(row=6, column=1,
        value="Meets, Winds, Tables and Indirects tab give summary statistics for the seating"
        )
    with open('../README', 'r') as readme_file:
        for i, line in enumerate(readme_file, start=8):
            readme_ws.cell(row=i, column=1, value=line.strip())
    for cell in readme_ws['A']:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    workbook.save('../seatings.xlsx')

    if write_final:
        write_meet_stats_to_csv(all_meet_hist, all_repeat3, all_repeat4,
                                '../stats/meet.csv')
        write_stats_to_csv(all_wind_hist, '../stats/winds.csv')
        write_stats_to_csv(all_table_hist, '../stats/tables.csv')
        write_stats_to_csv(all_indirect_hist, '../stats/indirects.csv')


def write_stats_to_worksheet(worksheet, data: Dict[str, Counter]) -> None:
    """
    Write statistical data to a worksheet.

    Args:
        worksheet: The worksheet to write to.
        data (Dict[str, Counter]): Dictionary containing statistical data.
    """
    max_index = max(len(hist) for hist in data.values() if hist)
    
    header = ['Configuration'] + list(range(max_index + 1))
    worksheet.append(header)
    
    for config, hist in data.items():
        row = [config] + [hist[i] for i in range(min(len(hist), max_index + 1))]
        row.extend([0] * (max_index + 1 - len(row)))
        worksheet.append(row)

    prettify_sheet(worksheet)
    

def prettify_sheet(worksheet):
    worksheet.conditional_formatting.add(
        f'B2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}',
        CellIsRule(operator='equal', formula=['0'], font=Font(color='808080'))
        )


def write_meet_stats_to_worksheet(worksheet, meet_hist: Dict[str, Counter],
                                  repeat3: Dict[str, int],
                                  repeat4: Dict[str, int]) -> None:
    """
    Write meeting statistics and repeat counts to a worksheet.

    Args:
        worksheet: The worksheet to write to.
        meet_hist (Dict[str, Counter]): Dictionary of meeting history data.
        repeat3 (Dict[str, int]): Dictionary of 3-player repeat counts.
        repeat4 (Dict[str, int]): Dictionary of 4-player repeat counts.
    """
    max_index = max(len(hist) for hist in meet_hist.values() if hist)
    
    header = ['Configuration'] + ['Repeat3', 'Repeat4'] + list(range(max_index + 1))
    worksheet.append(header)
    
    for config in meet_hist.keys():
        row = ([config] + [repeat3[config], repeat4[config]] +
               [meet_hist[config][i] for i in range(min(len(meet_hist[config]), max_index + 1))])
        row.extend([0] * (max_index + 1 - len(row)))
        worksheet.append(row)
    prettify_sheet(worksheet)


def write_stats_to_csv(data: Dict[str, Counter], filename: str) -> None:
    """
    Write statistical data to a CSV file.

    Args:
        data (Dict[str, Counter]): Dictionary containing statistical data.
        filename (str): Name of the output CSV file.

    Returns:
        None
    """
    max_index = max(len(hist) for hist in data.values() if hist)
    
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        header = ['Configuration'] + list(range(max_index + 1))
        writer.writerow(header)
        
        for config, hist in data.items():
            row = [config] + [hist[i] for i in range(
                min(len(hist), max_index + 1))]
            row.extend([0] * (max_index + 1 - len(row)))
            writer.writerow(row)


def write_meet_stats_to_csv(meet_hist: Dict[str, Counter],
                            repeat3: Dict[str, int],
                            repeat4: Dict[str, int],
                            filename: str) -> None:
    """
    Write meeting statistics and repeat counts to a CSV file.

    Args:
        meet_hist (Dict[str, Counter]): Dictionary of meeting history data.
        repeat3 (Dict[str, int]): Dictionary of 3-player repeat counts.
        repeat4 (Dict[str, int]): Dictionary of 4-player repeat counts.
        filename (str): Name of the output CSV file.

    Returns:
        None
    """
    max_index = max(len(hist) for hist in meet_hist.values() if hist)
    
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        header = (['Configuration'] + ['Repeat3', 'Repeat4'] +
                  list(range(max_index + 1)))
        writer.writerow(header)
        
        for config in meet_hist.keys():
            row = ([config] + [repeat3[config], repeat4[config]] +
                   [meet_hist[config][i] for i in range(
                    min(len(meet_hist[config]), max_index + 1))])
            row.extend([0] * (max_index + 1 - len(row)))
            writer.writerow(row)


if __name__ == "__main__":
    stat_all(write_final=True)